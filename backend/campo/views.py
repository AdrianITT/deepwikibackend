from django.shortcuts import render
from rest_framework import viewsets
from django.http import JsonResponse, HttpResponse
from .models import Preservador, Matriz, Clave, Contenedor, Parametro, Prioridad, CustodiaExterna, Muestra
from .models import RegistroCustodia, Filtro, PreservadorMuestra, EstadoCustodia
from .serializers import PreservadorSerializer, MatrizSerializer, ClaveSerializer, ContenedorSerializer, ParametroSerializer 
from .serializers import PrioridadSerializer, CustodiaExternaSerializer, MuestraSerializer, RegistroCustodiaSerializer
from .serializers import FiltroSerializer, PreservadorMuestraSerializer, EstadoCustodiaSerializer

class PreservadorViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Preservador.objects.all()
    serializer_class = PreservadorSerializer
    
class MatrizViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Matriz.objects.all()
    serializer_class = MatrizSerializer
    
class ClaveViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Clave.objects.all()
    serializer_class = ClaveSerializer
    
class ContenedorViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Contenedor.objects.all()
    serializer_class = ContenedorSerializer
    
class ParametroViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Parametro.objects.all()
    serializer_class = ParametroSerializer
    
class PrioridadViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Prioridad.objects.all()
    serializer_class = PrioridadSerializer

class EstadoCustodiaViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = EstadoCustodia.objects.all()
    serializer_class = EstadoCustodiaSerializer
    
class CustodiaExternaViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = CustodiaExterna.objects.all()
    serializer_class = CustodiaExternaSerializer
    
class MuestraViewSet(viewsets.ModelViewSet):    
    #permission_classes = [IsAuthenticated]
    queryset = Muestra.objects.all()
    serializer_class = MuestraSerializer
    
class RegistroCustodiaViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = RegistroCustodia.objects.all()
    serializer_class = RegistroCustodiaSerializer
    
class FiltroViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = Filtro.objects.all()
    serializer_class = FiltroSerializer
    
class PreservadorMuestraViewSet(viewsets.ModelViewSet):
    #permission_classes = [IsAuthenticated]
    queryset = PreservadorMuestra.objects.all()
    serializer_class = PreservadorMuestraSerializer
    
import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import CustodiaExterna, Muestra
import textwrap  # Para manejar el ajuste de líneas sin partir palabras
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side



def llenar_excel_con_informacion(request, custodia_id):
    # Cargar la plantilla existente
    ruta_plantilla = "media/cadenaCustodia/CustodiaExterna.xlsx"
    wb = openpyxl.load_workbook(ruta_plantilla)
    hoja = wb.active

    # Obtener la CustodiaExterna y datos relacionados usando las relaciones de Django
    custodia = get_object_or_404(CustodiaExterna, id=custodia_id)
    orden_trabajo = custodia.ordenTrabajo
    cotizacion = orden_trabajo.cotizacion
    cliente = cotizacion.cliente
    empresa = cliente.empresa
    receptor = custodia.receptor

    # Código de la orden de trabajo
    hoja['O4'] = orden_trabajo.codigo

    # Escribir el nombre de la empresa (F6 y B7 si es muy largo)
    nombre_empresa = empresa.nombre
    if len(nombre_empresa) <= 40:
        hoja['E6'] = nombre_empresa
    else:
        hoja['E6'] = nombre_empresa[:40]
        hoja['B7'] = nombre_empresa[40:]

    # Escribir la dirección completa en D8 y A9 si es muy larga
    direccion_empresa = f"{empresa.calle} {empresa.numero}, {empresa.colonia}, {empresa.ciudad}, {empresa.estado}, {empresa.codigoPostal}"
    celdas_direccion = ['D8', 'A9']
    ancho_celdas = [60, 60]  # Limite aproximado de caracteres por celda

    lineas = textwrap.wrap(direccion_empresa, width=ancho_celdas[0])
    hoja[celdas_direccion[0]] = lineas.pop(0) if lineas else ""
    for i in range(1, len(celdas_direccion)):
        texto_restante = " ".join(lineas)
        lineas = textwrap.wrap(texto_restante, width=ancho_celdas[i])
        hoja[celdas_direccion[i]] = lineas.pop(0) if lineas else ""

    # Escribir el nombre completo del cliente en la celda D10
    nombre_completo_cliente = f"{cliente.nombrePila} {cliente.apPaterno} {cliente.apMaterno}"
    hoja['D10'] = nombre_completo_cliente
    
    puesto_cliente = custodia.puestoCargoContacto
    hoja['O10'] = puesto_cliente if puesto_cliente else ""
    
    puntos_muestreo_autorizados = custodia.puntosMuestreoAutorizados
    hoja['H12'] = puntos_muestreo_autorizados
    
    muestreo_requerido = custodia.muestreoRequerido
    hoja['B16'] = muestreo_requerido
    
    fechaFinal = custodia.fechaFinal
    hoja['L15'] = fechaFinal.strftime("%d-%m-%Y") if fechaFinal else ""
    
    horaFinal = custodia.horaFinal
    hoja['L16'] = horaFinal if horaFinal else ""
    
    celular_cliente = cliente.celular
    hoja['D11'] = celular_cliente
    
    correo_cliente = cliente.correo
    hoja['L11'] = correo_cliente if correo_cliente else ""
    
    diagonal_side = Side(border_style="thin", color="000000")
    diagonal_border = Border(diagonal=diagonal_side, diagonalUp=True, diagonalDown=True)

    if custodia.modificacionOrdenTrabajo:
        celda = hoja["S12"]
    else:
        celda = hoja["T12"]

    celda.border = diagonal_border
    
        # Definir el estilo de la línea diagonal
    diagonal_side = Side(border_style="thin", color="000000")
    diagonal_border = Border(diagonal=diagonal_side, diagonalUp=True, diagonalDown=True)

    # Obtener el id de la prioridad y asignar la celda correspondiente
    if custodia.prioridad.id == 1:
        celda_prioridad = hoja["AS14"]
    elif custodia.prioridad.id == 2:
        celda_prioridad = hoja["AS15"]
    elif custodia.prioridad.id == 3:
        celda_prioridad = hoja["AS16"]
    else:
        celda_prioridad = None  # Por si acaso la prioridad tiene un id no esperado

    # Dibujar la "X" solo si se ha encontrado una celda válida
    if celda_prioridad:
        celda_prioridad.border = diagonal_border
        
    diagonal_side = Side(border_style="thin", color="000000")
    diagonal_border = Border(diagonal=diagonal_side, diagonalUp=True, diagonalDown=True)

    # Asignar la X según el valor de asesoriaGestionAmbiental
    if custodia.asesoriaGestionAmbiental:
        celda = hoja["AF48"]
    else:
        celda = hoja["AG48"]

    celda.border = diagonal_border


    # Obtener todas las muestras relacionadas con esta CustodiaExterna
    muestras = Muestra.objects.filter(custodiaExterna=custodia)
    fila_inicial = 20  # Fila inicial: B20, G20, J20, L20, P20, S20, etc.
    fila_final = 39    # Limitar a las filas hasta la 39

    # Diccionarios para almacenar las posiciones de cada muestra, parámetro, conservador y preservador
    parametros_usados = {}  # Almacena {clave_parametro: columna}
    columna_actual = 86  # Código ASCII para 'V'

    # Primero, escribir las muestras en las celdas básicas (B, G, J, L-O, P, S)
    for index, muestra in enumerate(muestras[:fila_final - fila_inicial + 1]):
        hoja[f"AP{fila_inicial + index}"] = muestra.numeroContenedor if muestra.numeroContenedor else ""
        hoja[f"AQ{fila_inicial + index}"] = muestra.origenMuestra if muestra.origenMuestra else ""
        hoja[f"B{fila_inicial + index}"] = muestra.identificacionCampo
        hoja[f"G{fila_inicial + index}"] = muestra.fechaMuestreo.strftime("%d-%m-%Y")
        hoja[f"J{fila_inicial + index}"] = muestra.horaMuestreo.strftime("%H:%M:%S")

        # Matriz
        codigo_matriz = muestra.matriz.codigo
        if codigo_matriz == "S":
            hoja[f"L{fila_inicial + index}"] = "X"
        elif codigo_matriz == "L":
            hoja[f"M{fila_inicial + index}"] = "X"
        elif codigo_matriz == "G":
            hoja[f"N{fila_inicial + index}"] = "X"
        elif codigo_matriz == "O":
            hoja[f"O{fila_inicial + index}"] = "X"

        # Volumen/Cantidad
        hoja[f"P{fila_inicial + index}"] = muestra.volumenCantidad if muestra.volumenCantidad else ""
        
        # Filtro
        hoja[f"S{fila_inicial + index}"] = muestra.filtro.codigo if muestra.filtro else ""

    columna_actual = 22
    
    # Segundo paso: Procesar parámetros, conservadores y preservadores
    for index, muestra in enumerate(muestras[:fila_final - fila_inicial + 1]):
        parametro = muestra.parametro
        conservador = muestra.contenedor
        preservadores_obj = list(muestra.preservador.all())
        preservador_ids = [p.id for p in preservadores_obj]

        clave_parametro = (parametro.id, conservador.codigo, tuple(sorted(preservador_ids)))

        if clave_parametro in parametros_usados:
            columna = parametros_usados[clave_parametro]
            hoja[f"{columna}{fila_inicial + index}"] = "X"
        else:
            columna = get_column_letter(columna_actual)  # Convierte correctamente el número a letra de columna
            parametros_usados[clave_parametro] = columna

            hoja[f"{columna}5"] = conservador.codigo  # Conservador en fila 5
            hoja[f"{columna}6"] = parametro.nombre      # Parámetro en fila 6

            preservadores_ids_str = ", ".join([str(p.id) for p in preservadores_obj])
            hoja[f"{columna}4"] = preservadores_ids_str

            hoja[f"{columna}{fila_inicial + index}"] = "X"
            columna_actual += 1
    
    nombre_completo_receptor = f"{receptor.nombrePila} {receptor.apPaterno} {receptor.apMaterno}"
    hoja['B41'] = nombre_completo_receptor
    
        # Lógica para manejar las observaciones
    observaciones = custodia.observaciones
    celdas_observaciones = ['E44', 'B45', 'B46']
    ancho_celdas = [95, 115, 115]  # Limite aproximado de caracteres por celda

    lineas = textwrap.wrap(observaciones, width=ancho_celdas[0])
    hoja[celdas_observaciones[0]] = lineas.pop(0) if lineas else ""
    for i in range(1, len(celdas_observaciones)):
        texto_restante = " ".join(lineas)
        lineas = textwrap.wrap(texto_restante, width=ancho_celdas[i])
        hoja[celdas_observaciones[i]] = lineas.pop(0) if lineas else ""
        
    observacionesModificacion = custodia.observacionesModificacion
    celdas_observacionesModificacion = ['E13', 'B14']
    ancho_celdas = [55, 65]
    
    lineas = textwrap.wrap(observacionesModificacion, width=ancho_celdas[0])
    hoja[celdas_observacionesModificacion[0]] = lineas.pop(0) if lineas else ""
    for i in range(1, len(celdas_observacionesModificacion)):
        texto_restante = " ".join(lineas)
        lineas = textwrap.wrap(texto_restante, width=ancho_celdas[i])
        hoja[celdas_observacionesModificacion[i]] = lineas.pop(0) if lineas else ""



    # Crear una respuesta HTTP para devolver el Excel modificado
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="CustodiaExterna_filled.xlsx"'
    wb.save(response)

    return response



def llenar_excel_con_informacion2(request, custodia_id):
    # Cargar la plantilla existente
    ruta_plantilla = "media/cadenaCustodia/CustodiaExterna.xlsx"
    wb = openpyxl.load_workbook(ruta_plantilla)
    hoja_original = wb.active
    hojas = [hoja_original]  # Lista para mantener todas las hojas
    
    # Obtener la CustodiaExterna y datos relacionados
    custodia = get_object_or_404(CustodiaExterna, id=custodia_id)
    orden_trabajo = custodia.ordenTrabajo
    cotizacion = orden_trabajo.cotizacion
    cliente = cotizacion.cliente
    empresa = cliente.empresa
    receptor = custodia.receptor

    # Obtener todas las muestras relacionadas
    muestras = Muestra.objects.filter(custodiaExterna=custodia)
    
    # Configuración de paginación
    muestras_por_hoja = 20  # De B20 a B39
    total_hojas_necesarias = (len(muestras) // muestras_por_hoja) + (1 if len(muestras) % muestras_por_hoja else 0)
    
    # Crear hojas adicionales si son necesarias
    for i in range(1, total_hojas_necesarias):
        nueva_hoja = wb.copy_worksheet(hoja_original)
        nueva_hoja.title = f"Hoja {i+1}"
        hojas.append(nueva_hoja)
        nueva_hoja.print_area = 'A1:AY50'

    
    # Procesar cada hoja con su conjunto de muestras
    for hoja_num, hoja in enumerate(hojas):
        inicio_muestras = hoja_num * muestras_por_hoja
        fin_muestras = inicio_muestras + muestras_por_hoja
        muestras_hoja = muestras[inicio_muestras:fin_muestras]

        # Copiar valores a las hojas adicionales
        if hoja_num > 0:  # No hacer esto en la primera hoja porque ya están llenas
            hoja['O4'] = hojas[0]['O4'].value
            hoja['E6'] = hojas[0]['E6'].value
            hoja['B7'] = hojas[0]['B7'].value
            hoja['D8'] = hojas[0]['D8'].value
            hoja['A9'] = hojas[0]['A9'].value
            hoja['D10'] = hojas[0]['D10'].value
            hoja['O10'] = hojas[0]['O10'].value
            hoja['D11'] = hojas[0]['D11'].value
            hoja['L11'] = hojas[0]['L11'].value
            hoja['H12'] = hojas[0]['H12'].value
            hoja['E13'] = hojas[0]['E13'].value
            hoja['B14'] = hojas[0]['B14'].value
            hoja['B16'] = hojas[0]['B16'].value
            hoja['L15'] = hojas[0]['L15'].value
            hoja['L16'] = hojas[0]['L16'].value
            hoja['B41'] = hojas[0]['B41'].value
            hoja['E44'] = hojas[0]['E44'].value
            hoja['B45'] = hojas[0]['B45'].value
            hoja['B46'] = hojas[0]['B46'].value

        # Solo escribir los datos comunes en la primera hoja
        if hoja_num == 0:
            # Escribir información común (código de orden, empresa, etc.)
            hoja['O4'] = orden_trabajo.codigo

            # Escribir el nombre de la empresa
            nombre_empresa = empresa.nombre
            if len(nombre_empresa) <= 40:
                hoja['E6'] = nombre_empresa
            else:
                hoja['E6'] = nombre_empresa[:40]
                hoja['B7'] = nombre_empresa[40:]

            # Escribir la dirección completa
            direccion_empresa = f"{empresa.calle} {empresa.numero}, {empresa.colonia}, {empresa.ciudad}, {empresa.estado}, {empresa.codigoPostal}"
            celdas_direccion = ['D8', 'A9']
            ancho_celdas = [60, 60]

            lineas = textwrap.wrap(direccion_empresa, width=ancho_celdas[0])
            hoja[celdas_direccion[0]] = lineas.pop(0) if lineas else ""
            for i in range(1, len(celdas_direccion)):
                texto_restante = " ".join(lineas)
                lineas = textwrap.wrap(texto_restante, width=ancho_celdas[i])
                hoja[celdas_direccion[i]] = lineas.pop(0) if lineas else ""

            # Escribir información del cliente
            hoja['D10'] = f"{cliente.nombrePila} {cliente.apPaterno} {cliente.apMaterno}"
            hoja['O10'] = custodia.puestoCargoContacto if custodia.puestoCargoContacto else ""
            hoja['H12'] = custodia.puntosMuestreoAutorizados
            hoja['B16'] = custodia.muestreoRequerido
            hoja['L15'] = custodia.fechaFinal.strftime("%d-%m-%Y") if custodia.fechaFinal else ""
            hoja['L16'] = custodia.horaFinal if custodia.horaFinal else ""
            hoja['D11'] = cliente.celular
            hoja['L11'] = cliente.correo if cliente.correo else ""
            
            # Marcas de verificación
            diagonal_side = Side(border_style="thin", color="000000")
            diagonal_border = Border(diagonal=diagonal_side, diagonalUp=True, diagonalDown=True)

            if custodia.modificacionOrdenTrabajo:
                celda = hoja["S12"]
            else:
                celda = hoja["T12"]
            celda.border = diagonal_border
            
            # Prioridad
            if custodia.prioridad.id == 1:
                celda_prioridad = hoja["AS14"]
            elif custodia.prioridad.id == 2:
                celda_prioridad = hoja["AS15"]
            elif custodia.prioridad.id == 3:
                celda_prioridad = hoja["AS16"]
            if celda_prioridad:
                celda_prioridad.border = diagonal_border
                
            if custodia.muestraCompuesta:
                hoja["AX41"].border = diagonal_border  # Dibuja diagonales en AX41 si es True

            if custodia.muestraPuntual:
                hoja["AX44"].border = diagonal_border  # Dibuja diagonales en AX44 si es True
                
            # Asesoría
            celda = hoja["AF48"] if custodia.asesoriaGestionAmbiental else hoja["AG48"]
            celda.border = diagonal_border

            # Observaciones
            observaciones = custodia.observaciones
            celdas_observaciones = ['E44', 'B45', 'B46']
            ancho_celdas = [95, 115, 115]

            lineas = textwrap.wrap(observaciones, width=ancho_celdas[0])
            hoja[celdas_observaciones[0]] = lineas.pop(0) if lineas else ""
            for i in range(1, len(celdas_observaciones)):
                texto_restante = " ".join(lineas)
                lineas = textwrap.wrap(texto_restante, width=ancho_celdas[i])
                hoja[celdas_observaciones[i]] = lineas.pop(0) if lineas else ""
                
            observacionesModificacion = custodia.observacionesModificacion
            celdas_observacionesModificacion = ['E13', 'B14']
            ancho_celdas = [55, 65]
            
            lineas = textwrap.wrap(observacionesModificacion, width=ancho_celdas[0])
            hoja[celdas_observacionesModificacion[0]] = lineas.pop(0) if lineas else ""
            for i in range(1, len(celdas_observacionesModificacion)):
                texto_restante = " ".join(lineas)
                lineas = textwrap.wrap(texto_restante, width=ancho_celdas[i])
                hoja[celdas_observacionesModificacion[i]] = lineas.pop(0) if lineas else ""

            # Nombre del receptor (solo en primera hoja)
            hoja['B41'] = f"{receptor.nombrePila} {receptor.apPaterno} {receptor.apMaterno}"

        # Procesar muestras para esta hoja
        fila_inicial = 20
        parametros_usados = {}
        columna_actual = 22  # Columna 'V'

        for index, muestra in enumerate(muestras_hoja):
            fila_actual = fila_inicial + index
            
            # Escribir datos básicos de la muestra
            hoja[f"AP{fila_actual}"] = muestra.numeroContenedor if muestra.numeroContenedor else ""
            hoja[f"AQ{fila_actual}"] = muestra.origenMuestra if muestra.origenMuestra else ""
            hoja[f"B{fila_actual}"] = muestra.identificacionCampo
            hoja[f"G{fila_actual}"] = muestra.fechaMuestreo.strftime("%d-%m-%Y")
            hoja[f"J{fila_actual}"] = muestra.horaMuestreo.strftime("%H:%M:%S")

            # Matriz
            codigo_matriz = muestra.matriz.codigo
            if codigo_matriz == "S":
                hoja[f"L{fila_actual}"] = "X"
            elif codigo_matriz == "L":
                hoja[f"M{fila_actual}"] = "X"
            elif codigo_matriz == "G":
                hoja[f"N{fila_actual}"] = "X"
            elif codigo_matriz == "O":
                hoja[f"O{fila_actual}"] = "X"

            # Volumen/Cantidad
            hoja[f"P{fila_actual}"] = muestra.volumenCantidad if muestra.volumenCantidad else ""
            
            # Filtro
            hoja[f"S{fila_actual}"] = muestra.filtro.codigo if muestra.filtro else ""

        # Procesar parámetros, conservadores y preservadores
        for index, muestra in enumerate(muestras_hoja):
            fila_actual = fila_inicial + index
            parametro = muestra.parametro
            conservador = muestra.contenedor
            preservadores_obj = list(muestra.preservador.all())
            preservador_ids = [p.id for p in preservadores_obj]

            clave_parametro = (parametro.id, conservador.codigo, tuple(sorted(preservador_ids)))

            if clave_parametro in parametros_usados:
                columna = parametros_usados[clave_parametro]
                hoja[f"{columna}{fila_actual}"] = "X"
            else:
                columna = get_column_letter(columna_actual)
                parametros_usados[clave_parametro] = columna

                hoja[f"{columna}5"] = conservador.codigo
                hoja[f"{columna}6"] = parametro.nombre
                hoja[f"{columna}4"] = ", ".join([str(p.id) for p in preservadores_obj])

                hoja[f"{columna}{fila_actual}"] = "X"
                columna_actual += 1
                
    # Crear la respuesta HTTP con el Excel modificado
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="CustodiaExterna_filled.xlsx"'
    wb.save(response)

    return response











def custodiaexternadata(request, custodia_id):
    custodia = get_object_or_404(CustodiaExterna, id=custodia_id)
    orden_trabajo = custodia.ordenTrabajo
    cotizacion = orden_trabajo.cotizacion
    cliente = cotizacion.cliente
    empresa = cliente.empresa

    data = {
        "ordenDeTrabajo": {
            "id": orden_trabajo.id,
            "codigo": orden_trabajo.codigo
        },
        "empresa": {
            "id": empresa.id,
            "nombre": empresa.nombre,
            "direccion": {
                "calle": empresa.calle,
                "numero": empresa.numero,
                "colonia": empresa.colonia,
                "ciudad": empresa.ciudad,
                "estado": empresa.estado,
                "codigoPostal": empresa.codigoPostal
            }
        },
        "cliente": {
            "id": cliente.id,
            "nombre": f"{cliente.nombrePila} {cliente.apPaterno} {cliente.apMaterno}",
            "celular": cliente.celular,
            "correo": cliente.correo
        },
        "custodiaExterna": {
            "id": custodia.id,
            "contacto": custodia.contacto,
            "puestoOCargoDelContacto": custodia.puestoCargoContacto or "",
            "celularDelContacto": custodia.celularContacto or "",
            "correoDelContacto": custodia.correoContacto or "",
            "puntosDeMuestreoAutorizados": custodia.puntosMuestreoAutorizados,
            "modificacionDeLaOrdenDeTrabajo": custodia.modificacionOrdenTrabajo,
            "observacionesDeLaModificacion": custodia.observacionesModificacion or "",
            "muestreoRequerido": custodia.muestreoRequerido,
            "fechaFinal": custodia.fechaFinal.strftime("%d-%m-%Y") if custodia.fechaFinal else "",
            "horaFinal": custodia.horaFinal.strftime("%H:%M:%S") if custodia.horaFinal else "",
            "prioridad": {
                "id": custodia.prioridad.id,
                "codigo": custodia.prioridad.codigo,
                "descripcion": custodia.prioridad.descripcion
            },
            "receptor": {
                "id": custodia.receptor.id,
                "nombre": f"{custodia.receptor.nombrePila} {custodia.receptor.apPaterno} {custodia.receptor.apMaterno}"
            },
            "observaciones": custodia.observaciones or "",
            "estado": {
                "id": custodia.estado.id,
                "descripcion": custodia.estado.descripcion
            },
            "solicitudDeAsesoriaEnGestionAmbiental": custodia.asesoriaGestionAmbiental,
            "muestraCompuesta": custodia.muestraCompuesta,
            "idDeLaMuestraCompuesta": custodia.idMuestraCompuesta or "",
            "muestraPuntual": custodia.muestraPuntual,
            "idDeLaMuestraPuntual": custodia.idMuestraPuntual or "",
            "muestras": [
                {
                    "id": muestra.id,
                    "identificacionDeCampo": muestra.identificacionCampo,
                    "fechaDeMuestreo": muestra.fechaMuestreo.strftime("%d-%m-%Y") if muestra.fechaMuestreo else "",
                    "horaDeMuestreo": muestra.horaMuestreo.strftime("%H:%M:%S") if muestra.horaMuestreo else "",
                    "matriz": {
                        "id": muestra.matriz.id,
                        "codigo": muestra.matriz.codigo,
                        "nombre": muestra.matriz.nombre
                    },
                    "volumenOCantidad": muestra.volumenCantidad or "",
                    "filtro": {
                        "id": muestra.filtro.id if muestra.filtro else None,
                        "codigo": muestra.filtro.codigo if muestra.filtro else "",
                        "descripcion": muestra.filtro.descripcion if muestra.filtro else ""
                    },
                    "parametro": {
                        "id": muestra.parametro.id,
                        "nombre": muestra.parametro.nombre
                    },
                    "conservador": {
                        "id": muestra.contenedor.id,
                        "codigo": muestra.contenedor.codigo,
                        "nombre": muestra.contenedor.nombre
                    },
                    "preservadores": [
                        {
                            "id": preservador.id,
                            "nombre": preservador.nombre
                        } for preservador in muestra.preservador.all()
                    ],
                    "numeroDeContenedor": muestra.numeroContenedor,
                    "origenDeLaMuestra": muestra.origenMuestra,
                } for muestra in custodia.muestra_set.all()
            ]
        }
    }

    return JsonResponse(data, safe=False)

def allcustodiaexterna(request):
    custodias = CustodiaExterna.objects.all()
    
    data = []
    for custodia in custodias:
        orden_trabajo = custodia.ordenTrabajo
        empresa = orden_trabajo.cotizacion.cliente.empresa
        prioridad = custodia.prioridad
        estado = custodia.estado
        
        custodia_data = {
            "custodiaExterna": {
                "id": custodia.id,
                "fechaFinal": custodia.fechaFinal.strftime("%d-%m-%Y") if custodia.fechaFinal else ""
            },
            "ordenTrabajo": {
                "id": orden_trabajo.id,
                "codigo": orden_trabajo.codigo
            },
            "empresa": {
                "id": empresa.id,
                "nombre": empresa.nombre
            },
            "prioridad": {
                "id": prioridad.id,
                "codigo": prioridad.codigo,
                "descripcion": prioridad.descripcion
            },
            "estado": {
                "id": estado.id,
                "descripcion": estado.descripcion
            }
        }
        
        data.append(custodia_data)
    
    return JsonResponse(data, safe=False)
